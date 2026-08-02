"""Leitura da matriz curricular oficial (``dados/matriz_curricular.xlsx``).

Esquema de abas adotado (documentado em ``docs/DICIONARIO_DADOS.md``):

1. ``Componentes``   — todos os componentes do curso (obrigatórios,
   optativos pré-aprovados, extensão, estágio, TCC, AAC...), diferenciados
   pelo campo ``tipo`` — não por prefixo/sufixo de código. Pré-requisitos e
   correquisitos são colunas desta própria aba (listas separadas por
   ``|``) — ver ``_lista_ids_pipe``/``_parse_prerequisitos``/
   ``_parse_correquisitos`` abaixo. Núcleo, áreas, temas transversais,
   conteúdos e competências de cada componente **não** são colunas daqui —
   são derivados na direção inversa, a partir da coluna ``componentes`` de
   cada aba de catálogo (item 3-7 abaixo).
2. ``Equivalencias`` — codigo_origem -> codigo_destino.
3. ``Nucleos``       — catálogo de núcleos curriculares (id, nome, descrição,
   componentes).
4. ``Areas``         — catálogo de áreas de formação (id, nome, descrição,
   componentes).
5. ``Temas``         — catálogo de temas transversais (id, nome, descrição,
   fonte normativa, status, componentes).
6. ``Conteudos``     — catálogo de conteúdos curriculares exigidos por
   alguma DCN (id, descrição, obrigatório, fonte, componentes).
7. ``Competencias``  — catálogo de competências do curso (id, descrição,
   obrigatória, fonte, componentes).
8. ``Bibliografia``  — catálogo de referências bibliográficas (chave, tipo,
   autor, título, ano...) — ver ``ppcgen.geradores.bibliografia``, que a
   renderiza em BibTeX/biblatex válido no momento da geração; não existe
   ``.bib`` estático em ``dados/``.
9. ``Legislacao``    — catálogo de referenciais legais do curso (id, nome,
   tipo, documento, ano, observações).
10. ``Certificacoes`` — opcional: certificacao_id -> codigo_componente (0+).
11. ``Autoridades``   — opcional: folha de rosto do PPC (reitor, vice-reitor,
    coordenador do curso...) — cargo, nome, observações, na ordem em que
    aparecem na planilha.
12. ``Comissao``      — opcional: membros da comissão de elaboração deste
    PPC (uma coluna ``membro``, texto livre, na ordem de exibição); o
    título da comissão é ``comissao.titulo`` na aba ``Perfil``, não vive
    aqui.
13. ``Perfil``        — chave/valor (``ler_configuracao_perfil`` abaixo):
    info/curso/instituição/currículo/oferta/capa/comissão/arquivos/geração/
    saída do perfil — não existe mais ``perfil.yaml``, esta aba concentra
    toda a configuração que não é curricular.

Em cada aba de catálogo (3-7), ``componentes`` é uma célula com códigos de
componente separados por ``|`` — os componentes vinculados àquele item. O
leitor monta o índice invertido no momento da carga: para cada código
listado, acrescenta o id do item de catálogo ao campo correspondente do
``ComponenteCurricular`` (``nucleo`` é cardinalidade 1 — primeiro núcleo a
reivindicar o componente vence; ``ppcgen.validadores.referenciais`` reporta
tanto códigos inexistentes quanto núcleos conflitantes). Um código listado
em ``componentes`` que não existe na aba ``Componentes`` não é erro do
leitor (Seção 29 — nunca perder dado silenciosamente): fica preservado em
``<Catalogo>.componentes`` (bruto) para o validador reportar.

Não há mais aba ``Curso``: a versão curricular é ``perfil.info.versao``
(Seção sobre fontes únicas em ``docs/DICIONARIO_DADOS.md``) — outras abas
que a planilha tenha (ex.: um fluxograma visual próprio de cada curso) não
fazem parte deste esquema e não são lidas por este módulo.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from ppcgen.excecoes import ArquivoNaoEncontrado, FormatoInvalido
from ppcgen.modelos import (
    AreaFormacao,
    Autoridade,
    CargaHoraria,
    Competencia,
    ComponenteCurricular,
    Conteudo,
    Correquisito,
    Curriculo,
    EntradaBibliografica,
    Equivalencia,
    NucleoCurricular,
    PreRequisito,
    ReferenciaisCurso,
    ReferencialCurricular,
    TemaTransversal,
    TipoComponente,
)

ABAS_OBRIGATORIAS = ("Componentes",)
ABAS_OPCIONAIS = (
    "Equivalencias",
    "Nucleos",
    "Areas",
    "Temas",
    "Conteudos",
    "Competencias",
    "Bibliografia",
    "Legislacao",
    "Certificacoes",
    "Autoridades",
    "Comissao",
)

_SUFIXO_OPCIONAL = "(opcional)"


def _linhas(planilha) -> list[dict]:
    """Converte uma planilha (com cabeçalho na 1ª linha) em lista de dicts."""

    linhas = list(planilha.iter_rows(values_only=True))
    if not linhas:
        return []
    cabecalho = [str(c).strip() if c is not None else "" for c in linhas[0]]
    resultado = []
    for linha in linhas[1:]:
        if all(v is None for v in linha):
            continue
        resultado.append(dict(zip(cabecalho, linha)))
    return resultado


def _bool(valor, padrao: bool = False) -> bool:
    if valor is None or valor == "":
        return padrao
    if isinstance(valor, bool):
        return valor
    return str(valor).strip().upper() in {"TRUE", "VERDADEIRO", "1", "SIM"}


def _int_ou_none(valor) -> int | None:
    if valor is None or valor == "":
        return None
    return int(valor)


def _periodo(valor) -> int | None:
    """Igual a ``_int_ou_none``, mas também aceita texto livre em volta do
    número (``5º Período``, ``5ºPeriodo``, ``Período 5``...) — usa o
    primeiro número encontrado na célula. Só para a coluna ``periodo``:
    outros campos numéricos (cht/chp/.../ano) continuam exigindo número
    puro, para não mascarar erro de digitação real neles."""

    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    texto = str(valor).strip()
    try:
        return int(texto)
    except ValueError:
        pass
    encontrado = re.search(r"\d+", texto)
    if encontrado is None:
        raise ValueError(f"não foi possível extrair um número de período de {valor!r}")
    return int(encontrado.group())


def _str_ou_vazio(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _tipo_componente(valor) -> TipoComponente:
    valor = _str_ou_vazio(valor).lower() or "disciplina"
    try:
        return TipoComponente(valor)
    except ValueError as exc:
        validos = ", ".join(t.value for t in TipoComponente)
        raise FormatoInvalido(f"Tipo de componente inválido: '{valor}'. Valores aceitos: {validos}") from exc


def _lista_ids_pipe(valor) -> list[str]:
    """Split de uma célula com itens separados por ``|`` — usada tanto para
    a coluna ``componentes`` das abas de catálogo quanto para
    ``pre_requisitos``/``correquisitos`` da aba ``Componentes``. Células
    vazias viram lista vazia, nunca ``[""]``."""

    texto = _str_ou_vazio(valor)
    if not texto:
        return []
    return [item.strip() for item in texto.split("|") if item.strip()]


def _extrair_opcional(item: str) -> tuple[str, bool]:
    """Separa o sufixo `` (opcional)`` (case-insensitive) de um item de
    pré-requisito/correquisito. Não usa ``?`` como marcador porque códigos
    de componente já podem legitimamente conter ``?`` (Seção sobre
    ``CODIGO_PROVISORIO`` em ``docs/VALIDACOES.md``)."""

    if item.lower().endswith(_SUFIXO_OPCIONAL):
        return item[: -len(_SUFIXO_OPCIONAL)].strip(), True
    return item, False


def _parse_correquisitos(valor) -> list[Correquisito]:
    correquisitos = []
    for item in _lista_ids_pipe(valor):
        codigo, opcional = _extrair_opcional(item)
        correquisitos.append(Correquisito(codigo=codigo, opcional=opcional))
    return correquisitos


def _parse_prerequisitos(valor) -> list[PreRequisito]:
    """Cada item é um código de componente (com sufixo opcional `` (opcional)``)
    ou uma exigência de carga horária mínima acumulada, escrita
    ``>=NNNh`` (nunca um código mágico como ``"*"`` — ver
    ``ppcgen.validadores.prerequisitos``). Exemplo de célula:
    ``CTR401|CTR203 (opcional)|>=1200h``."""

    pre_requisitos = []
    for item in _lista_ids_pipe(valor):
        carga = re.fullmatch(r"\s*(?:>=\s*)?(\d+)\s*(?:h|hora|horas)\s*", item, flags=re.IGNORECASE)
        if carga:
            pre_requisitos.append(PreRequisito(codigo="", carga_horaria_minima=int(carga.group(1))))
            continue
        codigo, opcional = _extrair_opcional(item)
        pre_requisitos.append(PreRequisito(codigo=codigo, opcional=opcional))
    return pre_requisitos


def carregar_registros_referenciais(wb) -> ReferenciaisCurso:
    """Lê os catálogos de núcleos/áreas/temas/conteúdos/competências/
    bibliografia/legislação das abas de registro da própria matriz
    (``Nucleos``/``Areas``/``Temas``/``Conteudos``/``Competencias``/
    ``Bibliografia``/``Legislacao``) — substitui os antigos
    ``referenciais/*.yaml``. Retorna um :class:`ReferenciaisCurso` com
    todos os campos preenchidos a partir da matriz — nada mais vem de
    fora dela. Não vincula aos componentes ainda — isso é
    ``_aplicar_vinculos_catalogo``, chamada por ``carregar_matriz`` depois
    que a aba ``Componentes`` também estiver carregada."""

    referenciais = ReferenciaisCurso()

    if "Nucleos" in wb.sheetnames:
        for row in _linhas(wb["Nucleos"]):
            id_ = _str_ou_vazio(row.get("id"))
            if id_:
                referenciais.nucleos.append(
                    NucleoCurricular(
                        id=id_,
                        nome=_str_ou_vazio(row.get("nome")),
                        descricao=_str_ou_vazio(row.get("descricao")),
                        componentes=_lista_ids_pipe(row.get("componentes")),
                    )
                )

    if "Areas" in wb.sheetnames:
        for row in _linhas(wb["Areas"]):
            id_ = _str_ou_vazio(row.get("id"))
            if id_:
                referenciais.areas.append(
                    AreaFormacao(
                        id=id_,
                        nome=_str_ou_vazio(row.get("nome")),
                        descricao=_str_ou_vazio(row.get("descricao")),
                        componentes=_lista_ids_pipe(row.get("componentes")),
                    )
                )

    if "Temas" in wb.sheetnames:
        for row in _linhas(wb["Temas"]):
            id_ = _str_ou_vazio(row.get("id"))
            if id_:
                referenciais.temas_transversais.append(
                    TemaTransversal(
                        id=id_,
                        nome=_str_ou_vazio(row.get("nome")),
                        descricao=_str_ou_vazio(row.get("descricao")),
                        fonte_normativa=_str_ou_vazio(row.get("fonte_normativa")),
                        status=_str_ou_vazio(row.get("status")) or "ativo",
                        componentes=_lista_ids_pipe(row.get("componentes")),
                    )
                )

    if "Conteudos" in wb.sheetnames:
        for row in _linhas(wb["Conteudos"]):
            id_ = _str_ou_vazio(row.get("id"))
            if id_:
                referenciais.conteudos.append(
                    Conteudo(
                        id=id_,
                        descricao=_str_ou_vazio(row.get("descricao")),
                        obrigatorio=_bool(row.get("obrigatorio")),
                        fonte=_str_ou_vazio(row.get("fonte")),
                        componentes=_lista_ids_pipe(row.get("componentes")),
                    )
                )

    if "Competencias" in wb.sheetnames:
        for row in _linhas(wb["Competencias"]):
            id_ = _str_ou_vazio(row.get("id"))
            if id_:
                referenciais.competencias.append(
                    Competencia(
                        id=id_,
                        descricao=_str_ou_vazio(row.get("descricao")),
                        obrigatoria=_bool(row.get("obrigatoria")),
                        fonte=_str_ou_vazio(row.get("fonte")),
                        componentes=_lista_ids_pipe(row.get("componentes")),
                    )
                )

    if "Bibliografia" in wb.sheetnames:
        for row in _linhas(wb["Bibliografia"]):
            chave = _str_ou_vazio(row.get("chave"))
            if chave:
                referenciais.bibliografia.append(
                    EntradaBibliografica(
                        chave=chave,
                        tipo=_str_ou_vazio(row.get("tipo")) or "misc",
                        autor=_str_ou_vazio(row.get("autor")),
                        titulo=_str_ou_vazio(row.get("titulo")),
                        ano=_str_ou_vazio(row.get("ano")),
                        mes=_str_ou_vazio(row.get("mes")),
                        dia=_str_ou_vazio(row.get("dia")),
                        endereco=_str_ou_vazio(row.get("endereco")),
                        editora=_str_ou_vazio(row.get("editora")),
                        organizacao=_str_ou_vazio(row.get("organizacao")),
                        instituicao=_str_ou_vazio(row.get("instituicao")),
                        edicao=_str_ou_vazio(row.get("edicao")),
                        serie=_str_ou_vazio(row.get("serie")),
                        doi=_str_ou_vazio(row.get("doi")),
                        paginas=_str_ou_vazio(row.get("paginas")),
                        url=_str_ou_vazio(row.get("url")),
                        nota=_str_ou_vazio(row.get("nota")),
                    )
                )

    if "Legislacao" in wb.sheetnames:
        for row in _linhas(wb["Legislacao"]):
            id_ = _str_ou_vazio(row.get("id"))
            if id_:
                referenciais.legislacao.append(
                    ReferencialCurricular(
                        id=id_,
                        nome=_str_ou_vazio(row.get("nome")),
                        tipo=_str_ou_vazio(row.get("tipo")),
                        documento=_str_ou_vazio(row.get("documento")),
                        ano=_int_ou_none(row.get("ano")),
                        observacoes=_str_ou_vazio(row.get("observacoes")),
                        url=_str_ou_vazio(row.get("url")),
                        chave_bibliografica=_str_ou_vazio(row.get("chave_bibliografica")),
                    )
                )

    if "Autoridades" in wb.sheetnames:
        for row in _linhas(wb["Autoridades"]):
            cargo = _str_ou_vazio(row.get("cargo"))
            if cargo:
                referenciais.autoridades.append(
                    Autoridade(
                        cargo=cargo,
                        nome=_str_ou_vazio(row.get("nome")),
                        observacoes=_str_ou_vazio(row.get("observacoes")),
                    )
                )

    if "Comissao" in wb.sheetnames:
        for row in _linhas(wb["Comissao"]):
            membro = _str_ou_vazio(row.get("membro"))
            if membro:
                referenciais.comissao_membros.append(membro)

    return referenciais


def _aplicar_vinculos_catalogo(
    componentes_por_codigo: dict[str, ComponenteCurricular], referenciais: ReferenciaisCurso
) -> list[str]:
    """Monta o índice invertido: para cada item de catálogo, para cada
    código em ``item.componentes``, preenche o campo correspondente do
    ``ComponenteCurricular``. Código que não existe na aba ``Componentes``
    não é alterado aqui (fica só em ``item.componentes``, bruto, para
    ``ppcgen.validadores.referenciais`` reportar) — este leitor nunca
    lança exceção nem descarta dado por uma referência quebrada (Seção 29).

    Núcleo é cardinalidade 1: o primeiro núcleo (na ordem da aba
    ``Nucleos``) a reivindicar um código vence; reivindicações
    conflitantes viram aviso aqui e ``NUCLEO_MULTIPLO_PARA_COMPONENTE``
    (erro) no validador, que varre ``referenciais.nucleos`` diretamente.
    """

    avisos: list[str] = []

    for nucleo in referenciais.nucleos:
        for codigo in nucleo.componentes:
            componente = componentes_por_codigo.get(codigo)
            if componente is None:
                continue
            if componente.nucleo is None:
                componente.nucleo = nucleo.id
            elif componente.nucleo != nucleo.id:
                avisos.append(
                    f"{codigo}: aparece em 'componentes' de mais de um núcleo "
                    f"('{componente.nucleo}' e '{nucleo.id}') — mantido "
                    f"'{componente.nucleo}'."
                )

    for area in referenciais.areas:
        for codigo in area.componentes:
            componente = componentes_por_codigo.get(codigo)
            if componente is not None:
                componente.areas.append(area.id)

    for tema in referenciais.temas_transversais:
        for codigo in tema.componentes:
            componente = componentes_por_codigo.get(codigo)
            if componente is not None:
                componente.temas_transversais.append(tema.id)

    for conteudo in referenciais.conteudos:
        for codigo in conteudo.componentes:
            componente = componentes_por_codigo.get(codigo)
            if componente is not None:
                componente.conteudos.append(conteudo.id)

    for competencia in referenciais.competencias:
        for codigo in competencia.componentes:
            componente = componentes_por_codigo.get(codigo)
            if componente is not None:
                componente.competencias.append(competencia.id)

    return avisos


def carregar_matriz(caminho: str | Path) -> tuple[Curriculo, ReferenciaisCurso, list[str]]:
    """Lê a matriz curricular e retorna ``(Curriculo, referenciais,
    avisos_leitura)``.

    ``Curriculo.versao`` vem em branco daqui — a versão curricular é
    ``perfil.info.versao`` (já existe na aba ``Perfil``; não duplicamos o
    dado numa aba ``Curso`` separada). Quem chama com um ``Perfil`` em mãos
    deve fazer ``curriculo.versao = perfil.info.versao`` depois de carregar.

    ``referenciais`` já vem completo — núcleos/áreas/temas/conteúdos/
    competências (vinculados aos componentes) e bibliografia/legislação
    (catálogos simples, sem vínculo por componente) — tudo das abas de
    registro da própria matriz; nada precisa ser preenchido a partir de
    fora dela depois.

    ``avisos_leitura`` registra situações que o leitor não deve "corrigir"
    silenciosamente (ex.: célula de status ativo/inativo em branco) — cabe ao
    validador (Seção 9) decidir a severidade de cada uma.
    """

    caminho = Path(caminho)
    if not caminho.exists():
        raise ArquivoNaoEncontrado(f"Matriz curricular não encontrada: {caminho}")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    for aba in ABAS_OBRIGATORIAS:
        if aba not in wb.sheetnames:
            raise FormatoInvalido(f"Aba obrigatória ausente na matriz: '{aba}'")

    avisos: list[str] = []

    # ``componentes`` preserva TODAS as linhas, mesmo com código repetido —
    # um dict aqui (codigo -> componente) apagaria silenciosamente a
    # primeira ocorrência de um código duplicado, exatamente o tipo de
    # perda silenciosa que este projeto proíbe (Seção 29). Duplicatas reais
    # são responsabilidade de ``CODIGO_DUPLICADO``
    # (``ppcgen.validadores.codigos``), não do leitor.
    componentes: list[ComponenteCurricular] = []
    for row in _linhas(wb["Componentes"]):
        codigo = _str_ou_vazio(row.get("codigo"))
        if not codigo:
            continue
        if row.get("ativo") is None or row.get("ativo") == "":
            avisos.append(
                f"{codigo}: coluna 'ativo' em branco na aba Componentes — assumido "
                "ativo=True; defina explicitamente TRUE/FALSE."
            )
        componente = ComponenteCurricular(
            codigo=codigo,
            nome=_str_ou_vazio(row.get("nome")),
            tipo=_tipo_componente(row.get("tipo")),
            carga_horaria=CargaHoraria(
                teorica=_int_ou_none(row.get("cht")),
                pratica=_int_ou_none(row.get("chp")),
                ead=_int_ou_none(row.get("chd")),
                extensao=_int_ou_none(row.get("che")),
                total=_int_ou_none(row.get("tot")),
            ),
            periodo=_periodo(row.get("periodo")),
            ativo=_bool(row.get("ativo"), padrao=True),
            obrigatorio=_bool(row.get("obrigatorio")),
            pre_requisitos=_parse_prerequisitos(row.get("pre_requisitos")),
            correquisitos=_parse_correquisitos(row.get("correquisitos")),
            observacoes=_str_ou_vazio(row.get("observacoes")),
        )
        componentes.append(componente)

    equivalencias = []
    if "Equivalencias" in wb.sheetnames:
        for row in _linhas(wb["Equivalencias"]):
            origem = _str_ou_vazio(row.get("codigo_origem"))
            destino = _str_ou_vazio(row.get("codigo_destino"))
            if origem and destino:
                equivalencias.append(
                    Equivalencia(
                        codigo_origem=origem,
                        codigo_destino=destino,
                        observacao=_str_ou_vazio(row.get("observacao")),
                    )
                )

    referenciais = carregar_registros_referenciais(wb)

    # Códigos duplicados resolvem para a última ocorrência (mesmo critério
    # de ``Curriculo.por_codigo()``) — degenerado por definição, mas
    # ``CODIGO_DUPLICADO`` (``ppcgen.validadores.codigos``) já cobre esse
    # caso separadamente.
    componentes_por_codigo = {c.codigo: c for c in componentes}
    avisos.extend(_aplicar_vinculos_catalogo(componentes_por_codigo, referenciais))

    curriculo = Curriculo(
        versao="",
        componentes=componentes,
        equivalencias=equivalencias,
    )
    return curriculo, referenciais, avisos


def _valor_chave_valor(bruto):
    """Coage o valor de uma célula da aba ``Perfil`` para o tipo Python mais
    apropriado, sem coluna de tipo separada: número/booleano nativos do
    Excel passam direto; texto ``TRUE``/``FALSE``/``VERDADEIRO``/``FALSO``
    vira ``bool``; texto numérico vira ``int``/``float``; o resto fica
    string. Célula vazia devolve ``None`` — quem chama descarta a chave
    inteira nesse caso, deixando o default da dataclass valer."""

    if bruto is None:
        return None
    if isinstance(bruto, bool):
        return bruto
    if isinstance(bruto, (int, float)):
        return bruto
    texto = str(bruto).strip()
    if texto == "":
        return None
    if texto.upper() in {"TRUE", "VERDADEIRO"}:
        return True
    if texto.upper() in {"FALSE", "FALSO"}:
        return False
    try:
        return int(texto)
    except ValueError:
        pass
    try:
        return float(texto)
    except ValueError:
        pass
    return texto


def ler_configuracao_perfil(caminho: str | Path) -> dict:
    """Lê a aba ``Perfil`` (chave/valor, uma linha por campo, chave no
    formato ``secao.campo`` — ex.: ``curso.numero_periodos``) e devolve um
    dict aninhado por seção (``{"curso": {"numero_periodos": 8, ...}, ...}``)
    — mesmo formato que o antigo ``perfil.yaml``, consumido por
    ``ppcgen.config.carregar_perfil``. Célula ``valor`` em branco omite a
    chave (o default da dataclass correspondente se aplica)."""

    caminho = Path(caminho)
    if not caminho.exists():
        raise ArquivoNaoEncontrado(f"Matriz curricular não encontrada: {caminho}")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    if "Perfil" not in wb.sheetnames:
        raise FormatoInvalido(f"Aba obrigatória ausente na matriz: 'Perfil' (em {caminho})")

    resultado: dict[str, dict] = {}
    chaves_lidas: set[str] = set()
    for numero_linha, row in enumerate(_linhas(wb["Perfil"]), start=2):
        chave = _str_ou_vazio(row.get("chave"))
        if not chave:
            continue
        if chave.count(".") != 1:
            raise FormatoInvalido(
                f"Chave inválida na aba 'Perfil', linha {numero_linha}: {chave!r}. "
                "Use exatamente o formato 'secao.campo'."
            )
        if chave in chaves_lidas:
            raise FormatoInvalido(f"Chave duplicada na aba 'Perfil', linha {numero_linha}: {chave!r}.")
        chaves_lidas.add(chave)
        secao, campo = chave.split(".", 1)
        valor = _valor_chave_valor(row.get("valor"))
        if valor is None:
            continue
        resultado.setdefault(secao, {})[campo] = valor
    return resultado
