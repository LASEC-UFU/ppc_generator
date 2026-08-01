"""Criação e clonagem de perfis (Seção 11).

Mantido separado da CLI para que a lógica de "quais arquivos um perfil novo
precisa" fique testável isoladamente.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

from ppcgen.config import CAPITULOS_PADRAO
from ppcgen.excecoes import ConfiguracaoInvalida

_PASTAS_FICHAS = ("obrigatorias", "optativas", "extensao", "tcc", "estagio", "complementares")
_PASTAS_ANEXOS = ("resolucoes", "pareceres", "outros")


def _linhas_perfil_padrao(perfil_id: str, nome: str) -> list[tuple[str, object]]:
    """Linhas chave/valor padrão de um perfil novo pra aba ``Perfil`` —
    mesmos defaults que antes viviam no template de ``perfil.yaml``."""

    return [
        ("perfil.id", perfil_id),
        ("perfil.nome", nome),
        ("perfil.status", "rascunho"),
        ("perfil.versao", ""),
        ("perfil.descricao", ""),
        ("perfil.extends", None),
        ("curso.nome", nome),
        ("curso.nome_curto", ""),
        ("curso.nome_mercadologico", ""),
        ("curso.enfase_curricular", ""),
        ("curso.titulacao_conferida", ""),
        ("curso.tempo_minimo_integralizacao", ""),
        ("curso.tempo_maximo_integralizacao", ""),
        ("curso.vagas_ofertadas", ""),
        ("curso.eixo_tecnologico", ""),
        ("curso.area_tecnologica", ""),
        ("curso.codigo_cine", ""),
        ("curso.sigla", ""),
        ("curso.grau", ""),
        ("curso.modalidade", ""),
        ("curso.turno", ""),
        ("curso.regime_academico", "Semestral"),
        ("curso.numero_periodos", 8),
        ("curso.campus", ""),
        ("curso.municipio", ""),
        ("curso.estado", ""),
        ("instituicao.nome", ""),
        ("instituicao.sigla", ""),
        ("instituicao.unidade_academica", ""),
        ("curriculo.carga_horaria_total", None),
        ("curriculo.carga_obrigatoria", None),
        ("curriculo.carga_optativa_minima", None),
        ("curriculo.carga_extensao", None),
        ("curriculo.carga_aac", None),
        ("curriculo.carga_estagio", None),
        ("curriculo.carga_tcc", None),
        ("curriculo.percentual_minimo_extensao", None),
        ("curriculo.percentual_maximo_ead", None),
        ("curriculo.carga_horaria_maxima_periodo", None),
        ("curriculo.periodo_minimo_tcc", None),
        ("curriculo.periodo_minimo_estagio", None),
        ("arquivos.textos", "textos"),
        ("arquivos.fichas", "fichas"),
        ("arquivos.figuras", "figuras"),
        ("arquivos.anexos", "anexos"),
        ("arquivos.frontmatter", "frontmatter"),
        ("arquivos.overrides", "overrides"),
        ("arquivos.capitulos", "|".join(CAPITULOS_PADRAO)),
        ("geracao.template", "padrao"),
        ("geracao.anexar_fichas", True),
        ("geracao.anexar_resolucoes", True),
        ("geracao.gerar_fluxo_curricular", True),
        ("geracao.gerar_representacao_grafica", True),
        ("geracao.gerar_relatorio_validacao", True),
        ("geracao.compilar_pdf", True),
        ("geracao.interromper_em_erro", True),
        ("saida.nome_base", f"PPC_{perfil_id}"),
        ("saida.gerar_corpo", True),
        ("saida.gerar_completo", True),
    ]


def _criar_matriz_vazia(caminho: Path, perfil_id: str, nome: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    perfil_sheet = wb.create_sheet("Perfil")
    perfil_sheet.append(["chave", "valor"])
    for chave, valor in _linhas_perfil_padrao(perfil_id, nome):
        perfil_sheet.append([chave, valor])

    componentes = wb.create_sheet("Componentes")
    componentes.append(
        [
            "codigo", "nome", "tipo", "periodo", "ativo", "obrigatorio",
            "cht", "chp", "chd", "che", "tot", "observacoes",
            "pre_requisitos", "correquisitos",
        ]
    )
    for aba, cabecalho in (
        ("Equivalencias", ["codigo_origem", "codigo_destino", "observacao"]),
        ("Nucleos", ["id", "nome", "descricao", "componentes"]),
        ("Areas", ["id", "nome", "descricao", "componentes"]),
        ("Temas", ["id", "nome", "descricao", "fonte_normativa", "status", "componentes"]),
        ("Conteudos", ["id", "descricao", "obrigatorio", "fonte", "componentes"]),
        ("Competencias", ["id", "descricao", "obrigatoria", "fonte", "componentes"]),
        (
            "Bibliografia",
            [
                "chave", "tipo", "autor", "titulo", "ano", "mes", "dia", "endereco",
                "editora", "organizacao", "instituicao", "edicao", "serie", "doi",
                "paginas", "url", "nota",
            ],
        ),
        ("Legislacao", ["id", "nome", "tipo", "documento", "ano", "observacoes", "url", "chave_bibliografica"]),
    ):
        ws = wb.create_sheet(aba)
        ws.append(cabecalho)

    wb.save(caminho)


def criar_perfil(pasta_perfis: Path, perfil_id: str, nome: str) -> Path:
    """Cria a estrutura inicial de um perfil novo e vazio em
    ``pasta_perfis/<perfil_id>/``."""

    destino = pasta_perfis / perfil_id
    if destino.exists():
        raise ConfiguracaoInvalida(f"Já existe um perfil em {destino}.")

    destino.mkdir(parents=True)
    (destino / "textos").mkdir()
    (destino / "frontmatter").mkdir()
    (destino / "figuras" / "imagens_capitulos").mkdir(parents=True)
    (destino / "figuras" / "diagramas").mkdir(parents=True)
    (destino / "overrides" / "latex").mkdir(parents=True)
    (destino / "overrides" / "estilos").mkdir(parents=True)
    for sub in _PASTAS_FICHAS:
        (destino / "fichas" / sub).mkdir(parents=True)
    for sub in _PASTAS_ANEXOS:
        (destino / "anexos" / sub).mkdir(parents=True)

    _criar_matriz_vazia(destino / "matriz_curricular.xlsx", perfil_id, nome)

    for nome_capitulo in CAPITULOS_PADRAO:
        titulo = nome_capitulo.replace("_", " ").title()
        (destino / "textos" / f"{nome_capitulo}.tex").write_text(
            f"% TODO: escrever o capítulo '{titulo}'.\n", encoding="utf-8"
        )

    (destino / "frontmatter" / "capa.yaml").write_text("capa:\n  ano: null\n", encoding="utf-8")
    (destino / "frontmatter" / "autoridades.yaml").write_text("autoridades: []\n", encoding="utf-8")
    (destino / "frontmatter" / "comissao.yaml").write_text(
        "comissao:\n  titulo: \"Equipe de elaboração deste Projeto Pedagógico\"\n  membros: []\n",
        encoding="utf-8",
    )

    (destino / "README.md").write_text(
        f"# Perfil: {perfil_id}\n\n{nome}\n\nStatus: rascunho — gerado por "
        "`python -m ppcgen perfil-criar`. Preencha `matriz_curricular.xlsx`: "
        "a aba `Perfil` (info/curso/instituição/currículo/oferta/geração/"
        "saída, uma linha chave/valor por campo); componentes; as abas "
        "Nucleos/Areas/Temas/Conteudos/Competencias, cada uma com sua coluna "
        "`componentes` listando os códigos vinculados, separados por `|`; a "
        "aba Bibliografia, uma linha por referência — o `.bib` compilado é "
        "sempre gerado a partir dela, nunca editado à mão; e a aba "
        "Legislacao, uma linha por referencial legal. Preencha também "
        "`textos/` antes de validar.\n",
        encoding="utf-8",
    )

    return destino


def clonar_perfil(
    pasta_perfis: Path, origem_id: str, destino_id: str, versao: str | None = None
) -> Path:
    """Clona um perfil existente: copia arquivos editáveis, atualiza id/versão,
    e registra a origem — nunca copia artefatos gerados (não existem dentro
    da pasta do perfil, por isolamento — Seção 12) nem caches."""

    origem = pasta_perfis / origem_id
    destino = pasta_perfis / destino_id
    if not origem.exists():
        raise ConfiguracaoInvalida(f"Perfil de origem '{origem_id}' não existe em {origem}.")
    if destino.exists():
        raise ConfiguracaoInvalida(f"Já existe um perfil em {destino}.")

    shutil.copytree(
        origem,
        destino,
        ignore=shutil.ignore_patterns("__pycache__", "*.pyc", ".DS_Store"),
    )

    caminho_matriz = next(destino.glob("matriz_curricular.xls*"), None)
    if caminho_matriz is None:
        raise ConfiguracaoInvalida(f"Nenhum matriz_curricular.xlsm/.xlsx encontrado em {destino}.")

    eh_xlsm = caminho_matriz.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(caminho_matriz, keep_vba=eh_xlsm)
    ws = wb["Perfil"]
    linhas_por_chave = {
        str(linha[0].value).strip(): linha[1] for linha in ws.iter_rows(min_row=2) if linha[0].value
    }

    def _definir(chave: str, valor: str) -> None:
        if chave in linhas_por_chave:
            linhas_por_chave[chave].value = valor
        else:
            ws.append([chave, valor])

    _definir("perfil.id", destino_id)
    _definir("perfil.origem_clonagem", origem_id)
    if versao is not None:
        _definir("perfil.versao", versao)

    wb.save(caminho_matriz)

    return destino
