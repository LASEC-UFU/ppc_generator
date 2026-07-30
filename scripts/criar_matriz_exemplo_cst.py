"""Script de bootstrap (executar uma única vez) que cria a matriz curricular
inicial/exemplo do Curso Superior de Tecnologia em Controle e Automação em
``dados/matriz_curricular.xlsx``.

Depois de gerado, o arquivo passa a ser a fonte editável oficial (Seção 3) —
este script não é chamado pelo pipeline do ppcgen e não deve ser executado
novamente sobre um arquivo já editado manualmente (ele SOBRESCREVE o
destino). Os dados abaixo são um levantamento inicial ilustrativo (Seção 24):
nomes de componentes, cargas horárias, pré-requisitos, unidades de oferta
etc. devem ser revisados e aprovados pelo NDE/Colegiado do curso antes de
qualquer uso oficial.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
DESTINO = RAIZ / "dados" / "matriz_curricular.xlsx"

# (codigo, nome, tipo, periodo, obrigatorio, cht, chp, chd, che, nucleo, area(s), unidade_oferta)
COMPONENTES = [
    ("MAT101", "Cálculo Diferencial e Integral I", "disciplina", 1, True, 60, 0, 0, 0, "BASICO", ["MATEMATICA"], "FAMAT"),
    ("MAT102", "Geometria Analítica e Álgebra Linear", "disciplina", 1, True, 60, 0, 0, 0, "BASICO", ["MATEMATICA"], "FAMAT"),
    ("FIS101", "Física: Mecânica", "disciplina", 1, True, 60, 0, 0, 0, "BASICO", ["FISICA"], "INFIS"),
    ("PRG101", "Lógica e Algoritmos de Programação", "disciplina", 1, True, 30, 30, 0, 0, "BASICO", ["COMPUTACAO"], "FACOM"),
    ("HUM101", "Comunicação Técnica e Metodologia Científica", "disciplina", 1, True, 30, 0, 0, 0, "HUMANISTICO", ["FORMACAO_HUMANA"], "FAGEN"),

    ("MAT201", "Cálculo Diferencial e Integral II", "disciplina", 2, True, 60, 0, 0, 0, "BASICO", ["MATEMATICA"], "FAMAT"),
    ("FIS201", "Física: Eletricidade e Magnetismo", "disciplina", 2, True, 60, 0, 0, 0, "BASICO", ["FISICA"], "INFIS"),
    ("PRG201", "Programação Orientada a Objetos", "disciplina", 2, True, 30, 30, 0, 0, "BASICO", ["COMPUTACAO"], "FACOM"),
    ("ELT201", "Circuitos Elétricos I", "disciplina", 2, True, 45, 15, 0, 0, "TECNOLOGICO", ["ELETRICIDADE"], "FEAUT"),
    ("EXT201", "Atividades Curriculares de Extensão I", "extensao", 2, True, 0, 0, 0, 45, "HUMANISTICO", ["EXTENSAO"], "FEAUT"),

    ("MAT301", "Equações Diferenciais e Cálculo Numérico", "disciplina", 3, True, 60, 0, 0, 0, "BASICO", ["MATEMATICA"], "FAMAT"),
    ("ELT301", "Circuitos Elétricos II", "disciplina", 3, True, 45, 15, 0, 0, "TECNOLOGICO", ["ELETRICIDADE"], "FEAUT"),
    ("ELN301", "Eletrônica Analógica", "disciplina", 3, True, 45, 15, 0, 0, "TECNOLOGICO", ["ELETRONICA"], "FEAUT"),
    ("SDI301", "Sistemas Digitais", "disciplina", 3, True, 30, 30, 0, 0, "TECNOLOGICO", ["SISTEMAS_DIGITAIS"], "FEAUT"),
    ("EST301", "Estatística Aplicada", "disciplina", 3, True, 45, 0, 0, 0, "BASICO", ["MATEMATICA"], "FAMAT"),
    ("EXT301", "Atividades Curriculares de Extensão II", "extensao", 3, True, 0, 0, 0, 45, "HUMANISTICO", ["EXTENSAO"], "FEAUT"),

    ("ELN401", "Eletrônica Digital e Microcontroladores", "disciplina", 4, True, 30, 30, 0, 0, "TECNOLOGICO", ["SISTEMAS_EMBARCADOS"], "FEAUT"),
    ("INS401", "Instrumentação Industrial", "disciplina", 4, True, 45, 15, 0, 0, "TECNOLOGICO", ["INSTRUMENTACAO"], "FEAUT"),
    ("CTR401", "Sistemas de Controle I", "disciplina", 4, True, 60, 0, 0, 0, "TECNOLOGICO", ["CONTROLE"], "FEAUT"),
    ("ACI401", "Acionamentos Elétricos", "disciplina", 4, True, 45, 15, 0, 0, "TECNOLOGICO", ["ACIONAMENTOS"], "FEAUT"),
    ("GES401", "Gestão e Empreendedorismo", "disciplina", 4, True, 30, 0, 0, 0, "HUMANISTICO", ["GESTAO"], "FAGEN"),
    ("EXT401", "Atividades Curriculares de Extensão III", "extensao", 4, True, 0, 0, 0, 60, "HUMANISTICO", ["EXTENSAO"], "FEAUT"),

    ("CTR501", "Sistemas de Controle II", "disciplina", 5, True, 45, 15, 0, 0, "TECNOLOGICO", ["CONTROLE"], "FEAUT"),
    ("CLP501", "Controladores Lógicos Programáveis", "disciplina", 5, True, 30, 30, 0, 0, "TECNOLOGICO", ["AUTOMACAO_INDUSTRIAL"], "FEAUT"),
    ("RED501", "Redes Industriais de Comunicação", "disciplina", 5, True, 45, 15, 0, 0, "TECNOLOGICO", ["REDES_INDUSTRIAIS"], "FEAUT"),
    ("SEG501", "Segurança em Sistemas Automatizados", "disciplina", 5, True, 45, 0, 0, 0, "TECNOLOGICO", ["SEGURANCA"], "FEAUT"),
    ("ETC501", "Ética, Legislação e Sustentabilidade", "disciplina", 5, True, 30, 0, 0, 0, "HUMANISTICO", ["FORMACAO_HUMANA"], "FADIR"),
    ("EXT501", "Atividades Curriculares de Extensão IV", "extensao", 5, True, 0, 0, 0, 75, "HUMANISTICO", ["EXTENSAO"], "FEAUT"),

    ("SUP601", "Sistemas Supervisórios e SCADA", "disciplina", 6, True, 30, 30, 0, 0, "TECNOLOGICO", ["AUTOMACAO_INDUSTRIAL"], "FEAUT"),
    ("PIN601", "Projeto Integrador em Automação Industrial", "projeto_integrador", 6, True, 0, 90, 0, 0, "TECNOLOGICO", ["AUTOMACAO_INDUSTRIAL"], "FEAUT"),
    ("EMB601", "Sistemas Embarcados Aplicados à Automação", "disciplina", 6, True, 30, 30, 0, 0, "TECNOLOGICO", ["SISTEMAS_EMBARCADOS"], "FEAUT"),

    ("STG001", "Estágio Supervisionado", "estagio", None, True, 0, 160, 0, 0, "TECNOLOGICO", ["AUTOMACAO_INDUSTRIAL"], "FEAUT"),
    ("TCC001", "Trabalho de Conclusão de Curso", "tcc", None, True, 0, 90, 0, 0, "TECNOLOGICO", ["AUTOMACAO_INDUSTRIAL"], "FEAUT"),
    ("AAC001", "Atividades Acadêmicas Complementares", "atividade_complementar", None, True, None, None, None, None, "COMPLEMENTAR", ["FORMACAO_HUMANA"], "FEAUT"),

    ("OPT001", "Robótica Industrial", "carga_optativa", None, False, 60, 0, 0, 0, "OPTATIVO", ["AUTOMACAO_INDUSTRIAL"], "FEAUT"),
    ("OPT002", "Eficiência Energética Industrial", "carga_optativa", None, False, 60, 0, 0, 0, "OPTATIVO", ["ACIONAMENTOS"], "FEAUT"),
    ("OPT003", "Visão Computacional Aplicada à Automação", "carga_optativa", None, False, 45, 15, 0, 0, "OPTATIVO", ["COMPUTACAO"], "FACOM"),
    ("OPT004", "Libras", "carga_optativa", None, False, 30, 30, 0, 0, "OPTATIVO", ["FORMACAO_HUMANA"], "FACED"),
    ("OPT005", "Redes Sem Fio Industriais (IIoT)", "carga_optativa", None, False, 45, 15, 0, 0, "OPTATIVO", ["REDES_INDUSTRIAIS"], "FEAUT"),
]

# AAC001 tem carga variável (atividades diversas do estudante); usamos TOT=60
# fixo aqui para fins de conferência com curriculo.carga_aac.
CARGA_TOTAL_FIXA = {"AAC001": 60}

# (codigo_componente, codigo_prerequisito ou "", opcional, carga_horaria_minima)
PREREQUISITOS = [
    ("MAT201", "MAT101", False, None),
    ("FIS201", "FIS101", False, None),
    ("PRG201", "PRG101", False, None),
    ("ELT201", "FIS101", False, None),
    ("MAT301", "MAT201", False, None),
    ("ELT301", "ELT201", False, None),
    ("ELN301", "ELT201", False, None),
    ("SDI301", "PRG101", False, None),
    ("EST301", "MAT102", False, None),
    ("EXT301", "EXT201", False, None),
    ("ELN401", "SDI301", False, None),
    ("INS401", "ELN301", False, None),
    ("CTR401", "MAT301", False, None),
    ("CTR401", "ELT301", False, None),
    ("ACI401", "ELT301", False, None),
    ("EXT401", "EXT301", False, None),
    ("CTR501", "CTR401", False, None),
    ("CLP501", "ELN401", False, None),
    ("CLP501", "INS401", False, None),
    ("RED501", "ELN401", False, None),
    ("SEG501", "INS401", False, None),
    ("EXT501", "EXT401", False, None),
    ("SUP601", "CLP501", False, None),
    ("PIN601", "CTR501", False, None),
    ("EMB601", "ELN401", False, None),
    ("STG001", "", False, 1200),
    ("TCC001", "", False, 1500),
]

CORREQUISITOS = [
    ("SEG501", "CLP501", False),
    ("PIN601", "SUP601", False),
]

TEMAS = [
    ("HUM101", "DIREITOS_HUMANOS"),
    ("HUM101", "LIBRAS"),
    ("ETC501", "EDUCACAO_AMBIENTAL"),
    ("ETC501", "RELACOES_ETNICO_RACIAIS"),
    ("OPT004", "LIBRAS"),
]

COMPETENCIAS = [
    ("CTR401", "PROJETAR_SISTEMAS_CONTROLE"),
    ("CTR501", "PROJETAR_SISTEMAS_CONTROLE"),
    ("CLP501", "PROGRAMAR_CLP_SUPERVISORIO"),
    ("SUP601", "PROGRAMAR_CLP_SUPERVISORIO"),
    ("INS401", "ESPECIFICAR_INSTRUMENTACAO"),
    ("ACI401", "PROJETAR_ACIONAMENTOS"),
    ("RED501", "INTEGRAR_REDES_INDUSTRIAIS"),
    ("EMB601", "DESENVOLVER_SISTEMAS_EMBARCADOS"),
    ("ELN401", "DESENVOLVER_SISTEMAS_EMBARCADOS"),
    ("SEG501", "APLICAR_SEGURANCA_INDUSTRIAL"),
    ("GES401", "GERIR_PROJETOS_AUTOMACAO"),
    ("PIN601", "GERIR_PROJETOS_AUTOMACAO"),
    ("ETC501", "ATUAR_ETICA_SUSTENTABILIDADE"),
]

EQUIVALENCIAS = [
    ("OPT003", "FACOM_VISAO01", "Disciplina equivalente ofertada pela Faculdade de Computação."),
]


def montar_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    curso = wb.create_sheet("Curso")
    curso.append(["campo", "valor"])
    curso.append(["versao_curricular", "2026-1-exemplo"])
    curso.append(["data_geracao", "2026-07-29"])
    curso.append(
        ["observacoes", "Matriz inicial/exemplo do CST em Controle e Automação."]
    )

    componentes = wb.create_sheet("Componentes")
    componentes.append(
        [
            "codigo", "nome", "tipo", "periodo", "ativo", "obrigatorio",
            "codigo_provisorio", "cht", "chp", "chd", "che", "tot",
            "nucleo_id", "unidade_oferta", "ementa", "observacoes",
        ]
    )
    for codigo, nome, tipo, periodo, obrigatorio, cht, chp, chd, che, nucleo, areas, unidade in COMPONENTES:
        if codigo in CARGA_TOTAL_FIXA:
            tot = CARGA_TOTAL_FIXA[codigo]
        else:
            tot = cht + chp + chd + che
        componentes.append(
            [codigo, nome, tipo, periodo, True, obrigatorio, False, cht, chp, chd, che, tot, nucleo, unidade, "", ""]
        )

    preq = wb.create_sheet("Pre-requisitos")
    preq.append(["codigo_componente", "codigo_prerequisito", "opcional", "carga_horaria_minima"])
    for row in PREREQUISITOS:
        preq.append(list(row))

    creq = wb.create_sheet("Correquisitos")
    creq.append(["codigo_componente", "codigo_correquisito", "opcional"])
    for row in CORREQUISITOS:
        creq.append(list(row))

    equiv = wb.create_sheet("Equivalencias")
    equiv.append(["codigo_origem", "codigo_destino", "observacao"])
    for row in EQUIVALENCIAS:
        equiv.append(list(row))

    areas_sheet = wb.create_sheet("Areas")
    areas_sheet.append(["codigo_componente", "area_id"])
    for c in COMPONENTES:
        codigo, area_list = c[0], c[10]
        for area_id in area_list:
            areas_sheet.append([codigo, area_id])

    temas_sheet = wb.create_sheet("Temas")
    temas_sheet.append(["codigo_componente", "tema_id"])
    for row in TEMAS:
        temas_sheet.append(list(row))

    competencias_sheet = wb.create_sheet("Competencias")
    competencias_sheet.append(["codigo_componente", "competencia_id"])
    for row in COMPETENCIAS:
        competencias_sheet.append(list(row))

    return wb


def main() -> None:
    if DESTINO.exists():
        raise SystemExit(
            f"{DESTINO} já existe — este script não sobrescreve dados possivelmente "
            "editados manualmente. Apague/mova o arquivo antes de rodar novamente, "
            "se realmente quiser recriá-lo do zero."
        )
    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb = montar_workbook()
    wb.save(DESTINO)
    print(f"Matriz curricular de exemplo criada em: {DESTINO}")


if __name__ == "__main__":
    main()
