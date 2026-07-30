"""Migra o PPC atual do curso de Engenharia de Computação (sistema legado:
``py/gen_docs.py`` + ``py/PPC_disciplinas_final.csv`` + ``include/*.tex``)
para o perfil ``dados/perfis/engenharia_computacao_2026_1/`` do novo sistema
(Seção 19).

Não apaga nem modifica nenhum arquivo legado — apenas lê e copia/converte
para o novo perfil. Idempotente o suficiente para revisão manual: recusa
sobrescrever um perfil já existente (mova/apague manualmente para
re-executar).

Uso:
    python scripts/migrar-perfil-legado.py
"""

from __future__ import annotations

import csv
import shutil

import openpyxl

from ppcgen.excecoes import ConfiguracaoInvalida
from ppcgen.leitores.csv import carregar_csv_legado
from ppcgen.modelos import TipoComponente
from ppcgen.utilitarios.caminhos import raiz_projeto

RAIZ = raiz_projeto()
CSV_LEGADO = RAIZ / "py" / "PPC_disciplinas_final.csv"
PERFIL_ID = "engenharia_computacao_2026_1"
PASTA_PERFIL = RAIZ / "dados" / "perfis" / PERFIL_ID

# Mesmas listas hardcoded em py/gen_docs.py (linhas ~1527-1531), extraídas
# aqui apenas para a migração — não ficam soltas no código do gerador.
CONTEUDOS_DCN_BASE = (
    "sistemas operacionais; compiladores; engenharia de software; interação "
    "humano-computador; redes de computadores; sistemas de tempo real; "
    "inteligência artificial e computacional; processamento de imagens; "
    "computação gráfica; banco de dados; dependabilidade; segurança; "
    "multimídia; sistemas embarcados; processamento paralelo; processamento "
    "distribuído; robótica; realidade virtual; automação; novos paradigmas "
    "de computação; matemática discreta; estruturas algébricas; matemática "
    "do contínuo; teoria dos grafos; análise combinatória; probabilidade e "
    "estatística; pesquisa operacional e otimização; teoria da computação; "
    "lógica; algoritmos e complexidade; linguagens formais e autômatos; "
    "abstração e estruturas de dados; fundamentos de linguagens; "
    "programação; modelagem computacional; métodos formais; análise, "
    "especificação, verificação e testes de sistemas; circuitos digitais; "
    "arquitetura e organização de computadores; avaliação de desempenho; "
    "ética e legislação; empreendedorismo; computação e sociedade; "
    "filosofia; metodologia cientifica; meio ambiente; fundamentos de "
    "administração; fundamentos de economia"
).split("; ")

CONTEUDOS_DCN_TEC = (
    "projeto de sistemas digitais; projeto de circuitos integrados; "
    "microeletrônica e nanoeletrônica; processamento digital de sinais; "
    "comunicação de dados; sistemas de controle; automação de projeto; "
    "transdutores; teoria dos semicondutores; teoria eletromagnética; "
    "eletrônica digital; eletrônica analógica; circuitos elétricos; "
    "eletricidade; física"
).split("; ")

MAPA_NUCLEO_NOME = {
    "BASICO": "Formação Básica",
    "HUMANISTICO": "Formação Humanística e de Extensão",
    "TECNOLOGICO": "Formação Tecnológica e Profissional",
    "COMPLEMENTAR": "Formação Optativa e Complementar",
}

MAPA_AREA_CC2020_NOME = {
    "CC2020_USUARIOS_ORGANIZACOES": "CC2020: Usuários e Organizações",
    "CC2020_MODELAGEM_SISTEMAS": "CC2020: Modelagem de Sistemas",
    "CC2020_ARQUITETURA_INFRAESTRUTURA": "CC2020: Arquitetura e Infraestrutura de Sistemas",
    "CC2020_DESENVOLVIMENTO_SOFTWARE": "CC2020: Desenvolvimento de Software",
    "CC2020_FUNDAMENTOS_SOFTWARE": "CC2020: Fundamentos de Software",
    "CC2020_HARDWARE": "CC2020: Hardware",
}

# Fallback de área quando o componente não tem nenhuma marcação CC2020 no
# CSV legado (a maioria dos componentes básicos/humanísticos) — decisão de
# migração explícita e documentada (docs/MIGRACAO.md), não uma inferência
# silenciosa: usa a mesma granularidade (grosseira) que já existia no
# núcleo, para satisfazer a exigência do novo modelo de que todo componente
# ativo tenha ao menos uma área.
MAPA_AREA_FALLBACK_POR_NUCLEO = {
    "BASICO": "FORMACAO_BASICA_GERAL",
    "HUMANISTICO": "FORMACAO_HUMANISTICA_GERAL",
    "TECNOLOGICO": "FORMACAO_TECNOLOGICA_GERAL",
    "COMPLEMENTAR": "FORMACAO_COMPLEMENTAR_GERAL",
}


def _ler_dcn_por_codigo() -> dict[str, tuple[list[int], list[int]]]:
    """Lê DCN_base/DCN_ecp diretamente do CSV (carregar_csv_legado não migra
    estas colunas — ver docs/MIGRACAO.md — então lemos aqui, isoladamente,
    só para alimentar referenciais/conteudos.yaml deste perfil)."""

    resultado: dict[str, tuple[list[int], list[int]]] = {}
    with open(CSV_LEGADO, encoding="utf-8", newline="") as f:
        for linha in csv.DictReader(f, delimiter=";"):
            codigo = (linha.get("Código") or "").strip()
            if not codigo:
                continue

            def _indices(campo: str) -> list[int]:
                bruto = (linha.get(campo) or "").strip()
                return [int(x) for x in bruto.split("/") if x.strip().lstrip("-").isdigit() and int(x) >= 0]

            resultado[codigo] = (_indices("DCN_base"), _indices("DCN_ecp"))
    return resultado


def migrar() -> None:
    if PASTA_PERFIL.exists():
        raise SystemExit(
            f"{PASTA_PERFIL} já existe — apague/mova manualmente antes de "
            "rodar a migração de novo (este script não sobrescreve)."
        )
    if not CSV_LEGADO.exists():
        raise ConfiguracaoInvalida(f"CSV legado não encontrado: {CSV_LEGADO}")

    resultado_import = carregar_csv_legado(CSV_LEGADO)
    curriculo = resultado_import.curriculo
    dcn_por_codigo = _ler_dcn_por_codigo()

    areas_usadas: set[str] = set()
    conteudos_usados: set[str] = set()

    for c in curriculo.componentes:
        if not c.areas and c.nucleo in MAPA_AREA_FALLBACK_POR_NUCLEO:
            c.areas = [MAPA_AREA_FALLBACK_POR_NUCLEO[c.nucleo]]
        areas_usadas.update(c.areas)

        dcn_base, dcn_tec = dcn_por_codigo.get(c.codigo, ([], []))
        for indice in dcn_base:
            if indice < len(CONTEUDOS_DCN_BASE):
                conteudo_id = f"DCN_BASE_{indice:02d}"
                c.conteudos.append(conteudo_id)
                conteudos_usados.add(conteudo_id)
        for indice in dcn_tec:
            if indice < len(CONTEUDOS_DCN_TEC):
                conteudo_id = f"DCN_TEC_{indice:02d}"
                c.conteudos.append(conteudo_id)
                conteudos_usados.add(conteudo_id)

    # --- estrutura de pastas -------------------------------------------------
    PASTA_PERFIL.mkdir(parents=True)
    (PASTA_PERFIL / "referenciais").mkdir()
    (PASTA_PERFIL / "textos").mkdir()
    (PASTA_PERFIL / "frontmatter").mkdir()
    (PASTA_PERFIL / "referencias").mkdir()
    (PASTA_PERFIL / "figuras").mkdir()
    for sub in ("obrigatorias", "optativas", "extensao", "tcc", "estagio", "complementares"):
        (PASTA_PERFIL / "fichas" / sub).mkdir(parents=True)
    for sub in ("resolucoes", "pareceres", "outros"):
        (PASTA_PERFIL / "anexos" / sub).mkdir(parents=True)
    (PASTA_PERFIL / "overrides" / "latex").mkdir(parents=True)
    (PASTA_PERFIL / "overrides" / "estilos").mkdir(parents=True)

    # --- matriz curricular ---------------------------------------------------
    _escrever_matriz(curriculo)

    # --- referenciais ---------------------------------------------------------
    _escrever_nucleos()
    _escrever_areas(areas_usadas)
    _escrever_conteudos(conteudos_usados)
    _escrever_legislacao()
    _escrever_temas_transversais()

    # --- figuras, bibliografia, fichas -----------------------------------------
    shutil.copyfile(RAIZ / "figure" / "auxiliary" / "ecp_logo.png", PASTA_PERFIL / "figuras" / "ecp_logo.png")
    shutil.copyfile(
        RAIZ / "include" / "backmatter" / "ppc2025.bib", PASTA_PERFIL / "referencias" / "bibliografia.bib"
    )
    _copiar_fichas(curriculo)

    print(f"Alertas de migração ({len(resultado_import.alertas_migracao)}):")
    for aviso in resultado_import.alertas_migracao:
        print("  -", aviso)
    print(f"\nPerfil '{PERFIL_ID}' criado em {PASTA_PERFIL}.")
    print(
        "Textos (identificacao.tex etc.), perfil.yaml e frontmatter/ ainda "
        "precisam ser copiados/adaptados de include/*.tex — ver "
        "docs/MIGRAR_PERFIL.md."
    )


def _escrever_matriz(curriculo) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    curso = wb.create_sheet("Curso")
    curso.append(["campo", "valor"])
    curso.append(["versao_curricular", curriculo.versao])
    curso.append(
        [
            "observacoes",
            "Migrado automaticamente de py/PPC_disciplinas_final.csv — ver docs/MIGRACAO.md.",
        ]
    )

    componentes = wb.create_sheet("Componentes")
    componentes.append(
        [
            "codigo", "nome", "tipo", "periodo", "ativo", "obrigatorio",
            "codigo_provisorio", "cht", "chp", "chd", "che", "tot",
            "nucleo_id", "unidade_oferta", "ementa", "observacoes",
        ]
    )
    preq_sheet = wb.create_sheet("Pre-requisitos")
    preq_sheet.append(["codigo_componente", "codigo_prerequisito", "opcional", "carga_horaria_minima"])
    creq_sheet = wb.create_sheet("Correquisitos")
    creq_sheet.append(["codigo_componente", "codigo_correquisito", "opcional"])
    areas_sheet = wb.create_sheet("Areas")
    areas_sheet.append(["codigo_componente", "area_id"])
    temas_sheet = wb.create_sheet("Temas")
    temas_sheet.append(["codigo_componente", "tema_id"])
    conteudos_sheet = wb.create_sheet("Conteudos")
    conteudos_sheet.append(["codigo_componente", "conteudo_id"])
    equiv_sheet = wb.create_sheet("Equivalencias")
    equiv_sheet.append(["codigo_origem", "codigo_destino", "observacao"])

    for c in curriculo.componentes:
        ch = c.carga_horaria
        componentes.append(
            [
                c.codigo, c.nome, c.tipo.value, c.periodo, c.ativo, c.obrigatorio,
                c.codigo_provisorio, ch.teorica, ch.pratica, ch.ead, ch.extensao, ch.total,
                c.nucleo, c.unidade_oferta, c.ementa, c.observacoes,
            ]
        )
        for preq in c.pre_requisitos:
            preq_sheet.append([c.codigo, preq.codigo, preq.opcional, preq.carga_horaria_minima])
        for creq in c.correquisitos:
            creq_sheet.append([c.codigo, creq.codigo, creq.opcional])
        for area_id in c.areas:
            areas_sheet.append([c.codigo, area_id])
        for tema_id in c.temas_transversais:
            temas_sheet.append([c.codigo, tema_id])
        for conteudo_id in c.conteudos:
            conteudos_sheet.append([c.codigo, conteudo_id])

    wb.save(PASTA_PERFIL / "matriz_curricular.xlsx")

    equivalencias_wb = openpyxl.Workbook()
    equivalencias_wb.active.append(["codigo_origem", "codigo_destino", "observacao"])
    equivalencias_wb.save(PASTA_PERFIL / "equivalencias.xlsx")


def _escrever_nucleos() -> None:
    linhas = "\n".join(f'  - id: {nid}\n    nome: "{nome}"' for nid, nome in MAPA_NUCLEO_NOME.items())
    conteudo = (
        "# Núcleos do curso de Engenharia de Computação, migrados da estrutura\n"
        "# de FORM_BAS/FORM_HUM/FORM_TEC/FORM_CMP do CSV legado.\n"
        f"nucleos:\n{linhas}\n"
    )
    (PASTA_PERFIL / "referenciais" / "nucleos.yaml").write_text(conteudo, encoding="utf-8")


def _escrever_areas(areas_usadas: set[str]) -> None:
    linhas = []
    for area_id in sorted(areas_usadas):
        nome = MAPA_AREA_CC2020_NOME.get(area_id)
        if nome is None:
            nome = area_id.replace("_", " ").title()
        linhas.append(f'  - id: {area_id}\n    nome: "{nome}"')
    conteudo = (
        "# Áreas migradas das marcações CC_UO1..CC_HW6 (CC2020) do CSV legado.\n"
        "# Componentes sem marcação CC2020 receberam uma área de fallback\n"
        "# derivada do núcleo (ver scripts/migrar-perfil-legado.py) — registrado\n"
        "# como decisão de migração em docs/MIGRACAO.md, não uma inferência\n"
        "# silenciosa.\n"
        "areas:\n" + "\n".join(linhas) + "\n"
    )
    (PASTA_PERFIL / "referenciais" / "areas_formacao.yaml").write_text(conteudo, encoding="utf-8")


def _escrever_conteudos(conteudos_usados: set[str]) -> None:
    linhas = []
    for i, descricao in enumerate(CONTEUDOS_DCN_BASE):
        cid = f"DCN_BASE_{i:02d}"
        if cid in conteudos_usados:
            linhas.append(
                f"  - id: {cid}\n    descricao: \"{descricao}\"\n    obrigatorio: true\n"
                "    fonte: MEC_CNE_CES_136_2012"
            )
    for i, descricao in enumerate(CONTEUDOS_DCN_TEC):
        cid = f"DCN_TEC_{i:02d}"
        if cid in conteudos_usados:
            linhas.append(
                f"  - id: {cid}\n    descricao: \"{descricao}\"\n    obrigatorio: true\n"
                "    fonte: MEC_CNE_CES_136_2012"
            )
    conteudo = (
        "# Conteúdos exigidos pelo Parecer CNE/CES nº 136/2012 (DCN Computação),\n"
        "# itens 3.1 (formação básica/tecnológica geral) e 3.3 (formação\n"
        "# tecnológica específica de Engenharia de Computação), migrados das\n"
        "# listas hardcoded `conteudos_curriculares`/`conteudos_basicos_tecnologicos`\n"
        "# de py/gen_docs.py. Específico deste perfil — não copiar para outros\n"
        "# cursos (Seção 24).\n"
        "conteudos:\n" + "\n".join(linhas) + "\n"
    )
    (PASTA_PERFIL / "referenciais" / "conteudos.yaml").write_text(conteudo, encoding="utf-8")


def _escrever_legislacao() -> None:
    conteudo = """# Referenciais legais ESPECÍFICOS deste perfil (Engenharia de Computação).
# Referenciais compartilháveis (extensão, educação ambiental, direitos
# humanos, libras, relações étnico-raciais) vêm de
# dados/compartilhados/legislacao/ via `heranca.legislacao` em perfil.yaml.
referenciais:
  - id: MEC_CNE_CES_136_2012
    nome: Diretrizes Curriculares Nacionais para os cursos de graduação em Computação
    tipo: parecer
    documento: Parecer CNE/CES nº 136/2012
    ano: 2012
    observacoes: >
      Base dos conteúdos obrigatórios em referenciais/conteudos.yaml deste
      perfil (itens 3.1 e 3.3).

  - id: MEC_CNE_CES_5_2016
    nome: Diretrizes Curriculares Nacionais para os cursos da área de Computação
    tipo: resolucao
    documento: Resolução CNE/CES nº 5, de 16 de novembro de 2016
    ano: 2016

  - id: ACM_CC2020
    nome: "Computing Curricula 2020: Paradigms for Global Computing Education"
    tipo: referencial_internacional
    documento: ACM/IEEE-CS Joint Task Force on Computing Curricula
    ano: 2020
    observacoes: >
      Referencial curricular internacional adotado voluntariamente por este
      curso (Seção 24: não é tratado como obrigatório para outros cursos).

  - id: CONFEA_380_1993
    nome: Título de Engenheiro de Computação
    tipo: resolucao
    documento: Resolução CONFEA nº 380, de 17 de dezembro de 1993
    ano: 1993
"""
    (PASTA_PERFIL / "referenciais" / "legislacao.yaml").write_text(conteudo, encoding="utf-8")


def _escrever_temas_transversais() -> None:
    conteudo = """# Temas transversais deste curso. Os identificadores coincidem com os
# usados no perfil de Controle e Automação (mesma base normativa nacional),
# mas cada perfil declara seu próprio arquivo — Seção 24 (não compartilhar
# implicitamente decisões de um curso com outro).
temas:
  - id: RELACOES_ETNICO_RACIAIS
    nome: Educação das Relações Étnico-Raciais
    fonte_normativa: Resolução CNE/CP nº 1/2004
    status: obrigatorio
  - id: LIBRAS
    nome: Ensino de Libras
    fonte_normativa: Decreto nº 5.626/2005
    status: obrigatorio
  - id: DIREITOS_HUMANOS
    nome: Educação em Direitos Humanos
    fonte_normativa: Resolução CNE nº 1/2012
    status: obrigatorio
  - id: EDUCACAO_AMBIENTAL
    nome: Educação Ambiental
    fonte_normativa: Resolução CNE nº 2/2012
    status: obrigatorio
  - id: PREVENCAO_DESASTRES
    nome: Educação para Prevenção e Redução de Desastres
    fonte_normativa: Lei nº 13.425/2017
    status: obrigatorio
"""
    (PASTA_PERFIL / "referenciais" / "temas_transversais.yaml").write_text(conteudo, encoding="utf-8")


def _copiar_fichas(curriculo) -> None:
    """Copia as fichas legadas (PDF) para as pastas do novo perfil,
    classificando pelo `tipo` do componente correspondente na matriz.

    A correspondência usa o leitor real de fichas (texto do PDF, não só o
    nome do arquivo — o nome sozinho é ambíguo nesta coleção legada: alguns
    arquivos usam o código no nome, outros o nome da disciplina, outros
    nenhum dos dois de forma padronizada). Fichas cujo código extraído não é
    reconhecido na matriz atual são casadas por nome (normalizado); se nem
    assim derem match, vão para `fichas/complementares/` com o nome
    original preservado, para revisão manual (não descartadas — Seção 15).
    """

    from ppcgen.leitores.fichas import carregar_fichas
    from ppcgen.utilitarios.textos import slug

    por_codigo = curriculo.por_codigo()
    por_nome = {slug(c.nome): c for c in curriculo.componentes}
    pastas_origem = [
        RAIZ / "fichas" / "SEI",
        RAIZ / "fichas" / "Fichas disciplinas 30h",
    ]

    mapa_tipo_pasta = {
        TipoComponente.EXTENSAO: "extensao",
        TipoComponente.ESTAGIO: "estagio",
        TipoComponente.TCC: "tcc",
        TipoComponente.ATIVIDADE_COMPLEMENTAR: "complementares",
        TipoComponente.CARGA_OPTATIVA: "optativas",
    }

    copiadas = 0
    nao_casadas = []
    for pasta in pastas_origem:
        if not pasta.exists():
            continue
        for ficha in carregar_fichas(pasta):
            arquivo = ficha.arquivo_origem
            componente = por_codigo.get(ficha.codigo)
            if componente is None and ficha.nome:
                componente = por_nome.get(slug(ficha.nome))
            if componente is None:
                # Último recurso: o nome do arquivo (sem extensão) bate com
                # o nome de alguma disciplina (caso de "Administração.pdf" etc.)
                componente = por_nome.get(slug(arquivo.stem))
            if componente is None:
                nao_casadas.append(arquivo.name)
            if componente is None:
                destino_sub = "complementares"
            elif componente.obrigatorio:
                destino_sub = "obrigatorias"
            else:
                destino_sub = mapa_tipo_pasta.get(componente.tipo, "complementares")
            shutil.copyfile(arquivo, PASTA_PERFIL / "fichas" / destino_sub / arquivo.name)
            copiadas += 1
    print(f"{copiadas} ficha(s) copiada(s) para dados/perfis/{PERFIL_ID}/fichas/.")
    if nao_casadas:
        print(
            f"{len(nao_casadas)} ficha(s) não puderam ser casadas com nenhum componente da "
            "matriz (foram para fichas/complementares/ para revisão manual):"
        )
        for nome in nao_casadas:
            print("  -", nome)


if __name__ == "__main__":
    migrar()
