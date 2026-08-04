from __future__ import annotations

import openpyxl
import pytest

from ppcgen.excecoes import FormatoInvalido
from ppcgen.leitores.excel import _parse_prerequisitos, carregar_matriz, ler_configuracao_perfil


def _matriz_minima(caminho):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    componentes = wb.create_sheet("Componentes")
    componentes.append(
        [
            "codigo",
            "nome",
            "tipo",
            "periodo",
            "ativo",
            "cht",
            "chp",
            "chd",
            "che",
            "tot",
            "observacoes",
            "pre_requisitos",
            "correquisitos",
        ]
    )
    componentes.append(["X1", "Disciplina X1", "disciplina", 1, True, 30, 0, 0, 0, 30, "", "", ""])
    componentes.append(["X2", "Disciplina X2", "disciplina", 2, True, 30, 0, 0, 0, 30, "", "X1", ""])

    nucleos = wb.create_sheet("Nucleos")
    nucleos.append(["id", "nome", "descricao", "componentes"])
    nucleos.append(["BASICO", "Formação Básica", "", "X1|X2"])

    areas = wb.create_sheet("Areas")
    areas.append(["id", "nome", "descricao", "componentes"])
    areas.append(["MATEMATICA", "Matemática", "", "X1|X2"])

    wb.save(caminho)


def test_carregar_matriz_basica(tmp_path):
    caminho = tmp_path / "matriz.xlsx"
    _matriz_minima(caminho)

    curriculo, referenciais, avisos = carregar_matriz(caminho)

    assert len(curriculo.componentes) == 2
    x2 = curriculo.por_codigo()["X2"]
    assert x2.pre_requisitos[0].codigo == "X1"
    assert x2.nucleo == "BASICO"
    assert x2.areas == ["MATEMATICA"]
    assert avisos == []
    assert referenciais.ids_nucleos() == {"BASICO"}
    assert referenciais.ids_areas() == {"MATEMATICA"}
    # legislação vem da aba Legislacao, igual às demais — não existe nesta
    # fixture mínima, então o catálogo fica vazio.
    assert referenciais.legislacao == []
    # competências vêm da aba Competencias, igual às demais — não existe
    # nesta fixture mínima, então o catálogo fica vazio.
    assert referenciais.competencias == []


def test_carregar_matriz_aba_obrigatoria_ausente(tmp_path):
    caminho = tmp_path / "matriz_invalida.xlsx"
    wb = openpyxl.Workbook()
    wb.save(caminho)

    with pytest.raises(FormatoInvalido):
        carregar_matriz(caminho)


def test_carregar_matriz_avisa_ativo_em_branco(tmp_path):
    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    componentes = wb.create_sheet("Componentes")
    componentes.append(
        ["codigo", "nome", "tipo", "periodo", "ativo", "cht", "chp", "chd", "che", "tot", "observacoes"]
    )
    componentes.append(["Y1", "Disciplina Y1", "disciplina", 1, None, 30, 0, 0, 0, 30, ""])
    wb.save(caminho)

    _curriculo, _referenciais, avisos = carregar_matriz(caminho)
    assert any("ativo" in aviso for aviso in avisos)


def test_curriculo_versao_vem_em_branco_da_matriz(tmp_path):
    """Não há mais aba ``Curso`` — a versão curricular é responsabilidade de
    quem chama (``perfil.info.versao``), nunca duplicada na planilha."""

    caminho = tmp_path / "matriz.xlsx"
    _matriz_minima(caminho)
    curriculo, _referenciais, _avisos = carregar_matriz(caminho)
    assert curriculo.versao == ""


def test_pre_requisitos_com_opcional_e_carga_horaria_minima(tmp_path):
    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    componentes = wb.create_sheet("Componentes")
    componentes.append(
        [
            "codigo",
            "nome",
            "tipo",
            "periodo",
            "ativo",
            "cht",
            "chp",
            "chd",
            "che",
            "tot",
            "observacoes",
            "pre_requisitos",
            "correquisitos",
        ]
    )
    componentes.append(
        ["Z1", "Estágio", "estagio", None, True, 0, 0, 0, 0, 300, "", "X1|X2 (opcional)|>=1200h", "X3 (opcional)"]
    )
    wb.save(caminho)

    curriculo, _referenciais, _avisos = carregar_matriz(caminho)
    z1 = curriculo.por_codigo()["Z1"]

    assert len(z1.pre_requisitos) == 3
    assert z1.pre_requisitos[0].codigo == "X1"
    assert z1.pre_requisitos[0].opcional is False
    assert z1.pre_requisitos[1].codigo == "X2"
    assert z1.pre_requisitos[1].opcional is True
    assert z1.pre_requisitos[2].codigo == ""
    assert z1.pre_requisitos[2].carga_horaria_minima == 1200

    assert len(z1.correquisitos) == 1
    assert z1.correquisitos[0].codigo == "X3"
    assert z1.correquisitos[0].opcional is True


@pytest.mark.parametrize("texto", ["1000 horas", "1800 hora", "1200h", ">= 900 H"])
def test_pre_requisito_de_carga_horaria_aceita_formatos_humanos(texto):
    requisito = _parse_prerequisitos(texto)[0]
    assert requisito.codigo == ""
    assert requisito.carga_horaria_minima is not None


def test_componentes_de_catalogo_vazio_nao_vincula_nada(tmp_path):
    """Célula ``componentes`` em branco numa aba de catálogo vira lista
    vazia (``_lista_ids_pipe``), nunca ``[""]`` — nenhum componente é
    vinculado."""

    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    componentes = wb.create_sheet("Componentes")
    componentes.append(["codigo", "nome", "tipo"])
    componentes.append(["W1", "Disciplina W1", "disciplina"])

    nucleos = wb.create_sheet("Nucleos")
    nucleos.append(["id", "nome", "descricao", "componentes"])
    nucleos.append(["BASICO", "Básico", "", None])

    areas = wb.create_sheet("Areas")
    areas.append(["id", "nome", "descricao", "componentes"])
    areas.append(["MATEMATICA", "Matemática", "", "  "])

    wb.save(caminho)

    curriculo, referenciais, _avisos = carregar_matriz(caminho)
    w1 = curriculo.por_codigo()["W1"]
    assert w1.nucleo is None
    assert w1.areas == []
    assert w1.temas_transversais == []
    assert w1.conteudos == []
    assert referenciais.nucleos[0].componentes == []
    assert referenciais.areas[0].componentes == []


def test_periodo_aceita_numero_puro_ou_texto_com_numero(tmp_path):
    """``periodo`` aceita tanto número puro quanto texto livre em volta
    (``5º Período``, ``5ºPeriodo``...) — usa o primeiro número encontrado.
    Outros campos numéricos (cht/chp/...) não ganham essa tolerância, só
    ``periodo``."""

    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    componentes = wb.create_sheet("Componentes")
    componentes.append(["codigo", "nome", "tipo", "periodo"])
    componentes.append(["X1", "Disciplina X1", "disciplina", 5])
    componentes.append(["X2", "Disciplina X2", "disciplina", "5º Período"])
    componentes.append(["X3", "Disciplina X3", "disciplina", "5ºPeriodo"])
    componentes.append(["X4", "Disciplina X4", "disciplina", "Período 5"])
    wb.save(caminho)

    curriculo, _referenciais, _avisos = carregar_matriz(caminho)
    por_codigo = curriculo.por_codigo()
    assert por_codigo["X1"].periodo == 5
    assert por_codigo["X2"].periodo == 5
    assert por_codigo["X3"].periodo == 5
    assert por_codigo["X4"].periodo == 5


def test_carregar_registros_referenciais_completo(tmp_path):
    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])

    nucleos = wb.create_sheet("Nucleos")
    nucleos.append(["id", "nome", "descricao"])
    nucleos.append(["BASICO", "Básico", "texto"])

    temas = wb.create_sheet("Temas")
    temas.append(["id", "nome", "descricao", "fonte_normativa", "status"])
    temas.append(["LIBRAS", "Libras", "", "Decreto nº 5.626/2005", "obrigatorio"])

    conteudos = wb.create_sheet("Conteudos")
    conteudos.append(["id", "descricao", "obrigatorio", "fonte"])
    conteudos.append(["DCN_01", "Conteúdo de teste", True, "DCN X"])

    competencias = wb.create_sheet("Competencias")
    competencias.append(["id", "descricao", "obrigatoria", "fonte"])
    competencias.append(["COMP_01", "Competência de teste", True, "Fonte X"])

    bibliografia = wb.create_sheet("Bibliografia")
    bibliografia.append(["chave", "tipo", "autor", "titulo", "ano", "url"])
    bibliografia.append(["teste_2024", "misc", "Autor Teste", "Título de Teste", "2024", "https://exemplo.org"])

    legislacao = wb.create_sheet("Legislacao")
    legislacao.append(["id", "nome", "tipo", "documento", "ano", "observacoes"])
    legislacao.append(["LDB", "Lei de Diretrizes e Bases", "lei", "Lei nº 9.394/1996", 1996, "texto"])

    autoridades = wb.create_sheet("Autoridades")
    autoridades.append(["cargo", "nome", "observacoes"])
    autoridades.append(["Reitor", "Fulano de Tal", ""])
    autoridades.append(["Coordenador(a) do Curso", "[a confirmar]", "PENDENTE: nome a confirmar"])

    comissao = wb.create_sheet("Comissao")
    comissao.append(["membro"])
    comissao.append(["Prof. Dr. Fulano -- presidente"])
    comissao.append(["Divisão de Projetos Pedagógicos"])

    wb.save(caminho)

    _curriculo, referenciais, _avisos = carregar_matriz(caminho)
    assert referenciais.nucleos[0].id == "BASICO"
    assert referenciais.temas_transversais[0].status == "obrigatorio"
    assert referenciais.conteudos[0].obrigatorio is True
    assert referenciais.competencias[0].id == "COMP_01"
    assert referenciais.competencias[0].obrigatoria is True
    assert referenciais.bibliografia[0].chave == "teste_2024"
    assert referenciais.bibliografia[0].tipo == "misc"
    assert referenciais.bibliografia[0].ano == "2024"
    assert referenciais.bibliografia[0].url == "https://exemplo.org"
    assert referenciais.legislacao[0].id == "LDB"
    assert referenciais.legislacao[0].documento == "Lei nº 9.394/1996"
    assert referenciais.legislacao[0].ano == 1996
    assert [a.cargo for a in referenciais.autoridades] == ["Reitor", "Coordenador(a) do Curso"]
    assert referenciais.autoridades[0].nome == "Fulano de Tal"
    assert referenciais.autoridades[1].observacoes == "PENDENTE: nome a confirmar"
    assert referenciais.comissao_membros == [
        "Prof. Dr. Fulano -- presidente",
        "Divisão de Projetos Pedagógicos",
    ]


def test_autoridades_e_comissao_ausentes_ficam_vazias(tmp_path):
    """As duas abas são opcionais (Seção 9) — matriz sem elas não deve
    quebrar nem inventar dado."""

    caminho = tmp_path / "matriz.xlsx"
    _matriz_minima(caminho)

    _curriculo, referenciais, _avisos = carregar_matriz(caminho)
    assert referenciais.autoridades == []
    assert referenciais.comissao_membros == []


def test_codigo_provisorio_e_unidade_oferta_derivados_do_codigo(tmp_path):
    """Nenhuma das duas colunas existe mais na planilha (Seção 9) — ambas
    são calculadas a partir de ``codigo``: ``unidade_oferta`` é o prefixo
    até o primeiro dígito/``!``; ``codigo_provisorio`` é ``True`` só para
    o prefixo ``FEELT!``."""

    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    componentes = wb.create_sheet("Componentes")
    componentes.append(["codigo", "nome", "tipo"])
    componentes.append(["FAMAT31011", "Cálculo I", "disciplina"])
    componentes.append(["FEELT!TDCA", "Tópicos em Automação", "disciplina"])
    wb.save(caminho)

    curriculo, _referenciais, _avisos = carregar_matriz(caminho)
    por_codigo = curriculo.por_codigo()

    oficial = por_codigo["FAMAT31011"]
    assert oficial.unidade_oferta == "FAMAT"
    assert oficial.codigo_provisorio is False

    provisorio = por_codigo["FEELT!TDCA"]
    assert provisorio.unidade_oferta == "FEELT"
    assert provisorio.codigo_provisorio is True


def test_vinculo_de_catalogo_a_componente_inexistente_nao_e_descartado(tmp_path):
    """Um código listado em ``componentes`` que não existe na aba
    Componentes não derruba dado nem lança exceção aqui (Seção 29) — fica
    preservado em ``NucleoCurricular.componentes`` (bruto) para
    ``ppcgen.validadores.referenciais`` reportar."""

    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    componentes = wb.create_sheet("Componentes")
    componentes.append(["codigo", "nome", "tipo", "ativo"])
    componentes.append(["V1", "Disciplina V1", "disciplina", True])

    nucleos = wb.create_sheet("Nucleos")
    nucleos.append(["id", "nome", "descricao", "componentes"])
    nucleos.append(["BASICO", "Básico", "", "V1|FANTASMA"])

    wb.save(caminho)

    curriculo, referenciais, avisos = carregar_matriz(caminho)
    v1 = curriculo.por_codigo()["V1"]
    assert v1.nucleo == "BASICO"
    assert referenciais.nucleos[0].componentes == ["V1", "FANTASMA"]


def test_ler_configuracao_perfil_coage_valores(tmp_path):
    """A aba ``Perfil`` (chave/valor) não tem coluna de tipo separada — a
    coerção é por heurística: número/booleano nativos passam direto, texto
    TRUE/FALSE vira bool, texto numérico vira int, célula em branco some do
    dict (deixa o default da dataclass valer), e uma chave desconhecida de
    ``instituicao.*`` continua disponível pra virar ``.extra`` depois."""

    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])

    perfil = wb.create_sheet("Perfil")
    perfil.append(["chave", "valor"])
    perfil.append(["perfil.id", "teste_coercao"])
    perfil.append(["curso.numero_periodos", 6])
    perfil.append(["geracao.compilar_pdf", "TRUE"])
    perfil.append(["geracao.anexar_fichas", False])
    perfil.append(["curriculo.percentual_minimo_extensao", "10"])
    perfil.append(["curriculo.carga_tcc", None])
    perfil.append(["instituicao.endereco", "Av. Exemplo, 123"])
    wb.save(caminho)

    bruto = ler_configuracao_perfil(caminho)

    assert bruto["perfil"]["id"] == "teste_coercao"
    assert bruto["curso"]["numero_periodos"] == 6
    assert bruto["geracao"]["compilar_pdf"] is True
    assert bruto["geracao"]["anexar_fichas"] is False
    assert bruto["curriculo"]["percentual_minimo_extensao"] == 10
    assert "carga_tcc" not in bruto["curriculo"]
    assert bruto["instituicao"]["endereco"] == "Av. Exemplo, 123"


def test_ler_configuracao_perfil_localiza_cabecalho_deslocado(tmp_path):
    """O cabeçalho ``chave``/``valor`` não precisa ser a 1ª linha — a aba
    ``Perfil`` aceita linhas extras acima dele (ex.: o indicador de
    validação de ``docs/indicador_validacao_vba.txt``, escrito direto nessa
    aba pelo VBA). Linhas antes do cabeçalho são ignoradas, e os números de
    linha usados nas mensagens de erro continuam apontando pra linha real."""

    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])

    perfil = wb.create_sheet("Perfil")
    perfil.append(["Verificação de regras (Python)", None])
    perfil.append(["●", "6 erro(s), 64 alerta(s) — verificado em 04/08/2026 06:01"])
    perfil.append([None, None])
    perfil.append(["chave", "valor"])
    perfil.append(["perfil.id", "teste_cabecalho_deslocado"])
    perfil.append(["curso.numero_periodos", 6])
    wb.save(caminho)

    bruto = ler_configuracao_perfil(caminho)

    assert bruto["perfil"]["id"] == "teste_cabecalho_deslocado"
    assert bruto["curso"]["numero_periodos"] == 6


def test_ler_configuracao_perfil_sem_cabecalho_gera_erro_formatado(tmp_path):
    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])

    perfil = wb.create_sheet("Perfil")
    perfil.append(["não é o cabeçalho esperado", None])
    perfil.append(["perfil.id", "nao_deveria_ser_lido"])
    wb.save(caminho)

    with pytest.raises(FormatoInvalido, match="cabeçalho"):
        ler_configuracao_perfil(caminho)


def test_carregar_perfil_le_arquivos_capitulos_customizado(tmp_path):
    """``arquivos.capitulos`` é o único campo de lista da aba ``Perfil`` —
    célula com itens separados por ``|`` vira ``list[str]``, na ordem
    declarada; sem a linha, cai no default (``CAPITULOS_PADRAO``)."""

    from ppcgen.config import CAPITULOS_PADRAO, carregar_perfil

    caminho = tmp_path / "matriz_curricular.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])

    perfil = wb.create_sheet("Perfil")
    perfil.append(["chave", "valor"])
    perfil.append(["perfil.id", "teste_capitulos"])
    perfil.append(["arquivos.capitulos", "introducao|conclusao"])
    wb.save(caminho)

    perfil_carregado = carregar_perfil(tmp_path, raiz_dados=tmp_path.parent)
    assert perfil_carregado.arquivos.capitulos == ["introducao", "conclusao"]

    caminho_sem_linha = tmp_path / "outro" / "matriz_curricular.xlsx"
    caminho_sem_linha.parent.mkdir()
    wb2 = openpyxl.Workbook()
    wb2.remove(wb2.active)
    wb2.create_sheet("Componentes").append(["codigo", "nome", "tipo"])
    perfil2 = wb2.create_sheet("Perfil")
    perfil2.append(["chave", "valor"])
    perfil2.append(["perfil.id", "teste_capitulos_default"])
    wb2.save(caminho_sem_linha)

    perfil_default = carregar_perfil(caminho_sem_linha.parent, raiz_dados=tmp_path.parent)
    assert perfil_default.arquivos.capitulos == list(CAPITULOS_PADRAO)


def test_carregar_perfil_coage_valor_numerico_em_campo_texto(tmp_path):
    """``curso.vagas_ofertadas`` é ``str`` (aceita texto livre, ex.: "30 no
    diurno, 20 no noturno"), mas uma célula puramente numérica na planilha
    (ex.: 30) chega da leitura do Excel como ``int``, não ``str`` — sem
    normalizar isso na borda de leitura (``_construir``), geradores que
    esperam string (ex.: ``ppcgen.utilitarios.latex.escapar``) quebram com
    ``TypeError`` ao tentar iterar um ``int``."""

    from ppcgen.config import carregar_perfil

    caminho = tmp_path / "matriz_curricular.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])

    perfil = wb.create_sheet("Perfil")
    perfil.append(["chave", "valor"])
    perfil.append(["perfil.id", "teste_vagas_numericas"])
    perfil.append(["curso.vagas_ofertadas", 30])
    wb.save(caminho)

    perfil_carregado = carregar_perfil(tmp_path, raiz_dados=tmp_path.parent)

    assert perfil_carregado.curso.vagas_ofertadas == "30"
    assert isinstance(perfil_carregado.curso.vagas_ofertadas, str)


def test_carregar_perfil_aba_ausente_gera_erro_formatado(tmp_path):
    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])
    wb.save(caminho)

    with pytest.raises(FormatoInvalido):
        ler_configuracao_perfil(caminho)


@pytest.mark.parametrize("chave", ["curso", "curso.numero.periodos"])
def test_ler_configuracao_perfil_rejeita_chave_malformada(tmp_path, chave):
    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Perfil"
    ws.append(["chave", "valor"])
    ws.append([chave, "valor"])
    wb.save(caminho)

    with pytest.raises(FormatoInvalido, match="Chave inválida"):
        ler_configuracao_perfil(caminho)


def test_ler_configuracao_perfil_rejeita_chave_duplicada(tmp_path):
    caminho = tmp_path / "matriz.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Perfil"
    ws.append(["chave", "valor"])
    ws.append(["perfil.id", "primeiro"])
    ws.append(["perfil.id", "segundo"])
    wb.save(caminho)

    with pytest.raises(FormatoInvalido, match="Chave duplicada"):
        ler_configuracao_perfil(caminho)


def test_carregar_perfil_le_aba_perfil_e_extrai_instituicao_extra(tmp_path):
    """Ponta a ponta via ``ppcgen.config.carregar_perfil``: chave
    ``instituicao.<desconhecida>`` (não é ``nome``/``sigla``/
    ``unidade_academica``) cai em ``InstituicaoConfig.extra``, e
    ``arquivos.matriz`` é sempre o nome do arquivo que foi aberto —
    autorreferente, nunca uma linha da aba."""

    from ppcgen.config import carregar_perfil

    caminho = tmp_path / "matriz_curricular.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Componentes").append(["codigo", "nome", "tipo"])

    perfil = wb.create_sheet("Perfil")
    perfil.append(["chave", "valor"])
    perfil.append(["perfil.id", "teste_extra"])
    perfil.append(["instituicao.nome", "Universidade de Teste"])
    perfil.append(["instituicao.site", "https://exemplo.org"])
    wb.save(caminho)

    perfil_carregado = carregar_perfil(tmp_path, raiz_dados=tmp_path.parent)

    assert perfil_carregado.info.id == "teste_extra"
    assert perfil_carregado.instituicao.nome == "Universidade de Teste"
    assert perfil_carregado.instituicao.extra["site"] == "https://exemplo.org"
    assert perfil_carregado.arquivos.matriz == "matriz_curricular.xlsx"


def test_carregar_perfil_rejeita_secao_desconhecida(tmp_path):
    from ppcgen.config import carregar_perfil
    from ppcgen.excecoes import ConfiguracaoInvalida

    caminho = tmp_path / "matriz_curricular.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Perfil"
    ws.append(["chave", "valor"])
    ws.append(["perfil.id", "teste"])
    ws.append(["seo_nao_suportada.valor", "x"])
    wb.save(caminho)

    with pytest.raises(ConfiguracaoInvalida, match="Seção.*desconhecida"):
        carregar_perfil(tmp_path, raiz_dados=tmp_path.parent)
